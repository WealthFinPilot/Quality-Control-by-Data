#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract.py — Extraction Excel (.xlsm) -> data/crm_data.json

Reproduit, en données ANONYMISÉES, le contenu métier des rapports de contrôle
de déviation d'un spectromètre d'émission optique.

Principe :
  - Source : feuille "Analyse" de chaque rapport (déjà calculée par le VBA d'origine).
  - On n'extrait QUE les données brutes nécessaires au recalcul côté navigateur :
    date, valeur mesurée par élément, écart-type des réplicats, valeur certifiée.
    Le moteur JS (src/engine.js) recalcule lui-même moyenne, σ, z-score et limites.
  - Tous les identifiants réels de standards sont remplacés par des alias neutres.

Le script est IDEMPOTENT : ré-exécutable sans créer de doublon (il réécrit le JSON).

Dépendance unique : openpyxl.

Cartographie de la feuille "Analyse" (identique dans les deux fichiers) :
  - Ligne 7            : en-tête. Colonne 1 = id standard ; puis, en alternance,
                         "<Élément>" (valeur) et "<Élément> z-score".
  - Lignes 8 .. (Average-1) : une mesure par date. Col 1 = date.
  - Ligne "Valeur certifiée" : valeur certifiée du CRM par élément.
  - Ligne "Lectures"   : "OK" = élément mesuré pour ce standard, "X" = non mesuré.
  - Bloc "SD <date>"   : écart-type des 3 réplicats, une ligne par mesure, même ordre.
"""

import json
import os
from datetime import datetime, date

import openpyxl

# --- Configuration ----------------------------------------------------------

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

# Fichiers source (jamais commités — voir .gitignore).
SOURCE_FILES = [
    os.path.join(ROOT, "Rapport-1.xlsm"),
    os.path.join(ROOT, "Rapport-2.xlsm"),
]

OUTPUT = os.path.join(ROOT, "data", "crm_data.json")

# Anonymisation : id réel du standard -> alias neutre.
# La correspondance contient des identifiants RÉELS : elle vit dans un fichier
# LOCAL gitignoré (`scripts/aliases.local.json`), JAMAIS dans le dépôt public.
# Voir `scripts/aliases.example.json` pour le format.
ALIASES_FILE = os.path.join(HERE, "aliases.local.json")


def load_aliases():
    """Charge la correspondance identifiant réel -> alias depuis le fichier local."""
    if not os.path.exists(ALIASES_FILE):
        raise SystemExit(
            f"[ERREUR] {os.path.relpath(ALIASES_FILE, ROOT)} introuvable. "
            f"Copiez scripts/aliases.example.json en scripts/aliases.local.json "
            f"et renseignez la correspondance réelle (fichier gitignoré)."
        )
    with open(ALIASES_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return {k: v for k, v in data.items() if not k.startswith("_")}


STANDARD_ALIASES = load_aliases()

REPLICATES = 3  # "Moyenne sur 3 Mesures" dans l'outil d'origine.


def as_iso(value):
    """Convertit une cellule date Excel en chaîne ISO 8601 (YYYY-MM-DD)."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def num(value):
    """Retourne un float si la cellule est numérique, sinon None."""
    if isinstance(value, (int, float)):
        return float(value)
    return None


def find_label_row(ws, label):
    """Numéro de la première ligne dont la colonne 1 vaut exactement `label`."""
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == label:
            return r
    return None


def extract_standard(path):
    """Extrait un standard depuis un fichier .xlsm. Retourne un dict ou None."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Analyse"]

    # En-tête ligne 7 : repérer les colonnes "valeur" (non z-score).
    header_row = 7
    element_cols = {}  # nom élément -> index colonne
    for c in range(2, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h is None:
            continue
        h = str(h).strip()
        if h and "z-score" not in h:
            element_cols[h] = c

    raw_id = str(ws.cell(header_row, 1).value).strip()
    alias = STANDARD_ALIASES.get(raw_id)
    if alias is None:
        raise SystemExit(
            f"[ERREUR] Standard '{raw_id}' sans alias dans STANDARD_ALIASES. "
            f"Refus d'écrire un identifiant réel."
        )

    avg_row = find_label_row(ws, "Average")
    cert_row = find_label_row(ws, "Valeur certifiée")
    lect_row = find_label_row(ws, "Lectures")
    # Première ligne du bloc des écarts-types de réplicats ("SD <date>").
    sd_first = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip().startswith("SD "):
            sd_first = r
            break

    # Lignes de mesures : de 8 à Average-1.
    measure_rows = list(range(8, avg_row))
    n = len(measure_rows)

    # Éléments réellement mesurés pour ce standard (Lectures == "OK").
    measured = []
    for name, col in element_cols.items():
        flag = str(ws.cell(lect_row, col).value).strip().upper()
        cert = num(ws.cell(cert_row, col).value)
        if flag == "OK" and cert is not None:
            measured.append((name, col, cert))

    certified_values = {name: cert for name, col, cert in measured}

    measurements = []
    for i, mr in enumerate(measure_rows):
        sd_r = sd_first + i  # alignement 1:1 mesure <-> écart-type réplicats
        values = {}
        std_by_element = {}
        for name, col, cert in measured:
            v = num(ws.cell(mr, col).value)
            if v is None:
                continue
            values[name] = round(v, 6)
            sd = num(ws.cell(sd_r, col).value)
            std_by_element[name] = round(sd, 6) if sd is not None else 0.0
        measurements.append({
            "date": as_iso(ws.cell(mr, 1).value),
            "values": values,
            "std_by_element": std_by_element,
            "replicates": REPLICATES,
        })

    wb.close()

    print(f"  {os.path.basename(path)} : {raw_id} -> {alias} | "
          f"{len(measured)} éléments mesurés, {n} mesures")

    return {
        "id": alias,
        "certified_values": certified_values,
        "measurements": measurements,
    }


def main():
    print("Extraction des standards (anonymisée) :")
    standards = [extract_standard(p) for p in SOURCE_FILES]
    standards.sort(key=lambda s: s["id"])

    payload = {
        "_comment": (
            "Données anonymisées. Identifiants de standards remplacés par des "
            "alias neutres (STD-A, STD-B). Généré par scripts/extract.py."
        ),
        "standards": standards,
    }

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    total = sum(len(s["measurements"]) for s in standards)
    print(f"OK -> {os.path.relpath(OUTPUT, ROOT)} "
          f"({len(standards)} standards, {total} mesures au total)")


if __name__ == "__main__":
    main()
